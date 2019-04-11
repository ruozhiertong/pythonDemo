#! /usr/bin/python
# -*- coding:utf-8 -*-

import openpyxl as xl

# 功能：按照某一列，将该列中相同单元格的行筛选到新的的表中。即先按照该列分组，然后将每组放到新的一张表中。

#@profile
def convert():
	wb = xl.load_workbook(filename= "test.xlsx",read_only=True)

	newwb = xl.Workbook()

	print("newsheets' names:" + str(newwb.sheetnames))


	sheet = wb["Sheet"]

	print(sheet.title)

	newSheets = {}

	for index ,row in enumerate(sheet.iter_rows()):
		# print(str(row[0].value))
		# print("cell value type is:")
		# print(type(row[0].value))
		# print(type(row[0]))
		if (index < 3) and row[0].value is None: #单元格 的元祖
			continue

		if str(row[0].value) not in newSheets.keys():
			newSheets[str(row[0].value)] = newwb.create_sheet(str(row[0].value))
			#print("newSheets' name" + str(newwb.sheetnames))
			#填充第二行表头。 
			for col in range(1,sheet.max_column + 1):
				newSheets[str(row[0].value)].cell(1, col).value = sheet.cell(2,col).value

		newSheetRow = newSheets[str(row[0].value)].max_row + 1
		for col , cell in enumerate(row):
			newSheets[str(row[0].value)].cell(newSheetRow, col+1).value = cell.value
			#print(cell.value)
		#print(type(row))
		#print(row)
		

	newwb.save("test_new.xlsx")


convert()



