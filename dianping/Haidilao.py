#!/usr/bin/python
# -*- coding:utf-8 -*-
import urllib
from urllib import request
from urllib import parse
from urllib.request import urlopen
import re
import json
import time
import random
import threading
import openpyxl as xl
from openpyxl.styles import Font
import sys

sys.path.append('../email')
import sendmail

sys.path.append('../log')
import log

#两者都可以。上面是具体的模块。下面是以包形式的。
# sys.path.append('..')
# import email_package.sendmail

#因为模块名比较长 ，可以as. 调用的时候 直接sm.sendMail 而不用 email_package.sendmail.senMail
# import email_package.sendmail as sm


#全局变量

#弃用了。 海底捞升级网站了。orz，刚做完就升级系统了
#url = 'http://www.haidilao.com/service/domestic'
#全国城市
china_city_url = 'http://www.haidilao.com/js/jquery.cityselect.js'
#海底捞门店城市
haidilao_city_url = 'http://www.haidilao.com/js/kuCity.js?v=8'



china_file = 'china.dat'

#注意这里全局list 如果要进行赋值的话， 最好设置成None。
#如果 china_cities =[] ,然后 china_cities = json.load(text）这样直接赋值的话，只是局部的，不能作用到全局的这个变量
china_cities = None
#查询到市级别
end_level = 2

# accept: text/plain, */*; q=0.01
# accept-encoding: gzip, deflate, br
# accept-language: zh-CN,zh;q=0.9,en;q=0.8
# easysitetoken: 
# origin: https://www.haidilao.com
# referer: https://www.haidilao.com/en/fwzx/mdss/index.html
# request-by: ajax-request-tag
# user-agent: Mozilla/5.0 (Macintosh; Intel Mac OS X 10_11_1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/74.0.3729.169 Safari/537.36
# x-requested-with: XMLHttpRequest
#https://www.haidilao.com/eportal/ui?moduleId=5&pageId=9c8cf76c4ca84fc686ca11aaa936f5c7&struts.portlet.action=/portlet/map-portlet!getMapDotDataByRegion.action&random=0.5606347888283192
#https://www.haidilao.com/eportal/ui?moduleId=5&pageId=181bb70fe3cc4c2596b3d7415178b749&struts.portlet.action=/portlet/map-portlet!getMapDotDataByRegion.action&random=0.621823985892598
#https://www.haidilao.com/eportal/ui?moduleId=5&pageId=181bb70fe3cc4c2596b3d7415178b749&struts.portlet.action=/portlet/map-portlet!getMapDotDataByRegion.action&random=0.12608788658106063
#https://www.haidilao.com/eportal/ui?moduleId=5&pageId=181bb70fe3cc4c2596b3d7415178b749&struts.portlet.action=/portlet/map-portlet!getMapDotDataByRegion.action&random=0.640995342813512
url = 'https://www.haidilao.com/eportal/ui?moduleId=5&pageId=181bb70fe3cc4c2596b3d7415178b749&struts.portlet.action=/portlet/map-portlet!getMapDotDataByRegion.action'

header = {
		  	'user-agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_11_1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/74.0.3729.169 Safari/537.36',
			'content-type': 'application/x-www-form-urlencoded; charset=UTF-8',
			'origin': 'https://www.haidilao.com',
		  	'referer': 'https://www.haidilao.com/en/fwzx/mdss/index.html',
		  	'x-requested-with': 'XMLHttpRequest',
		  	'request-by': 'ajax-request-tag',
		  	'asysiteparamkey':'', 
			'easysitetoken': '',
			'accept': 'text/plain, */*; q=0.01',
			'accept-encoding': 'gzip, deflate, br',
			'accept-language': 'zh-CN,zh;q=0.9,en;q=0.8'
		  	}


#海底捞门店搜索就是坑。 并不是完全按照行政区进行编排的。 搜索是按照详细地址中关键字匹配的。关键是其详细地址并没有严格按照行政区，如 厦门市xxx， 而你搜索福建省厦门市就没有结果。
#所以搜索结果实际上是有误差的。
#如果要想精确的话，可能要用美团，点评网这些数据吧。




execel_name = 'Haidilao.xlsx'
data_name = 'Haidilao.data'
email_list = ['1043096262@qq.com','songzehua2015@163.com']

city_num = 0
statics_data = {} # {"福建":{"num":12,"city":[{"福州":10},{"厦门":10}]}}



url_lock = threading.Lock()
city_lock = threading.Lock()

def exportExcel(data):
	wb = xl.Workbook() #xl.load_workbook(filename='result.xlsx')
	ws1 = wb.create_sheet("Hua")           #创建一个sheet
	ws1["A1"]='城市'
	ws1['B1']='门店'

	total = 0 
	row = 4
	# {"福建":{"num":12,"city":[{"福州":10},{"厦门":10}]}}
	for ele in data:
		ws1['A'+str(row)].font = Font(bold=True)
		ws1['A'+str(row)] = ele[0]
		num = ele[1].get("num")
		total += num 
		ws1['B'+str(row)].font = Font(bold=True)
		ws1['B'+str(row)] = num
		row +=1
		city  = ele[1].get('city') #列表
		for one in city: #one是dict
			for p,n in one.items():
				ws1['A'+str(row)] = p
				ws1['B'+str(row)] = n
				row +=1
		row +=1

	ws1['A2'].font = Font(bold=True)
	ws1['B2'].font = Font(bold=True)
	ws1['A2']='total'
	ws1['B2']=total

	wb.save(execel_name)

#cities 是一个dict。 root_name 就是省名
def visit_cities(city_name,city,root_name):

	if city is None:
		return

	if city['level'] > end_level:
		return
	if city['level'] < end_level:
		if city['child'] is not None:
			for x in city['child']:
				visit_cities(city['name'], x, root_name)
	elif city['level'] == end_level:
		#因为是随时变化的，而不是像header那样，放在局部，为了线程安全。
		data = {
			'queryContent': '',
			'country':'f919da82d3a0497e8cd06d9ae383f477',
			'myLat':'',
			'myLng':'' 
		}
		if city['name'] == '市辖区':
			data['queryContent'] = city_name
		else:
			data['queryContent'] = city['name']
		print(data)
		
		request_url = url + '&random=' + str(random.random())

		# print(request_url)
		text = getHtmlAdvance(request_url ,header = header, method = 'POST', query_str = '', data = data)

		if text is None:
			return


		json_data = json.loads(text)
		#print(json_data)
		shop_num = json_data.get('totalCount')
		if shop_num != 0:
			statics_data[root_name]['num'] += shop_num
			statics_data[root_name]['city'].append({data['queryContent']: shop_num})


# pro_idx 省份索引 china_cities
def process_province_request(thread_name, pro_idx):
	print(thread_name, pro_idx)
	province_dict = china_cities[pro_idx]
	#print(province_dict)
	pro_name = province_dict['name']
	statics_data[pro_name] = {}
	statics_data[pro_name]['num'] = 0
	statics_data[pro_name]['city'] = []

	#print(thread_name,pro_name, province_dict)
	visit_cities(pro_name, province_dict, pro_name)




# 为线程定义一个函数
def process_thread(threadName, shop_city, china_city):
	global city_num
	#print(threadName,"start")
	shop_city_num = len(shop_city)
	print(city_num, shop_city_num)
	while(True):
		#print(threadName,"run", city_num)
		url_lock.acquire()
		if(city_num >= shop_city_num):
			url_lock.release()
			break
		city = shop_city[city_num]
		city_num += 1
		url_lock.release()
		cityname = city[:city.find('|')]

		shop_num = getShopNum(cityname)
		citydict = {}
		citydict[cityname] = shop_num
		province = cityOfProvince(cityname,china_city)

		city_lock.acquire()
		#print(province)
		if(not statics_data.get(province)):
			statics_data[province] = {}
		if(not statics_data[province].get('num')):
			statics_data[province]['num'] = 0
		statics_data[province]['num'] += shop_num
		if(not statics_data[province].get('city')):
			statics_data[province]['city'] = []
		statics_data[province]['city'].append(citydict)

		city_lock.release()



def getHtml(url):

	fUrl = urlopen(url)
	text_b = fUrl.read()
	#print(text_b)
	charset = 'utf-8'
	index = text_b.find("charset".encode())
	fUrl.close()
	if "GBK".encode() in text_b[index:index + 30] or "gbk".encode() in text_b[index:index + 30]:
		charset = 'gbk'
	elif "utf-8".encode() in text_b[index:index + 30] or "UTF-8".encode() in text_b[index:index + 30]:
		charset = 'utf-8'

	text = text_b.decode(charset,'ignore')

	text = text.replace(' ','')
	text = text.replace('\'','"') #替换单引号，防止json处理时出错。
	text = text.replace('\t','')
	text = text.replace('\r\n','')
	#print(text)

	# with open("city.txt","w+") as f:
	# 	f.write(text)
	return text

# query_str '?key=value&key2=value2'
def getHtmlAdvance(url ,header = {}, method = 'GET', query_str = '', data = None):

	url = url + query_str

	if method == 'POST':
		data = parse.urlencode(data).encode('utf-8')
	req = request.Request(url , headers = header, data = data)

	flag = True
	count = 0
	fUrl = {}
	while flag and count < 3:
		try:
			fUrl = request.urlopen(req,timeout = 15)
			if fUrl.code == 200:
				flag = False
			count += 1
		except Exception as e:
			log.logInfo(str(e) + ',' + url)
			count += 1
	if flag:
		return None

	text_b = fUrl.read()
	fUrl.close()

	charset = 'utf-8'

	charset_str = fUrl.getheader('content-type')

	charset = None
	if charset_str is not None:
		if "gbk" in charset_str.lower():
			charset = 'gbk'
		elif "utf-8" in charset_str.lower():
			charset = 'utf-8'
		elif 'gb2312' in charset_str.lower():
			charset = 'gb2312'
	if charset is None:
		charset = 'utf-8'
		index = text_b.find("charset".encode())
		if "GBK".encode() in text_b[index:index + 30] or "gbk".encode() in text_b[index:index + 30]:
			charset = 'gbk'
		elif "utf-8".encode() in text_b[index:index + 30] or "UTF-8".encode() in text_b[index:index + 30]:
			charset = 'utf-8'
		elif "gb2312".encode() in text_b[index:index + 30] or "GB2312".encode() in text_b[index:index + 30]:
			charset = 'gb2312'
	text = text_b.decode(charset,'ignore')
	text = text.replace(' ','')
	text = text.replace('\'','"') #替换单引号，防止json处理时出错。
	text = text.replace('\t','')
	text = text.replace('\r\n','')

	time.sleep(5)
	return text

def getCity(html):
	#res = re.search('pro_city_data(.|\n)*"国外"\n}]\n}$',html)
	#print(html)
	#pattern = re.compile(r'pro_city_data.*"国外"\r\n}]\r\n}', re.M|re.S|re.DOTALL)
	pattern = re.compile(r'pro_city_data.*"国外"}]}', re.M|re.S|re.DOTALL)

	#print(pattern.findall(html))
	res = pattern.findall(html)
	#print(res[0])
	# print(len(res))

	# if res:
	# 	print(res.group())
	# 	print(res.group(1))
	# else:
	# 	print("Nothing found!!")
	pro_city_data = json.loads(res[0][res[0].find("{"):])
	# print(pro_city_data)
	return pro_city_data

def getHaidilaoCity(html):
	pattern = re.compile('varallCities=.*gx"]', re.M|re.S|re.DOTALL)
	res = pattern.findall(html)
	#print(res[0][res[0].find("["):])
	#print(len(res))
	shop_city_data = json.loads(res[0][res[0].find("["):])
	return shop_city_data

def getDataFromHtml(html, regex, start):
	#多行匹配re.M re.DOTALL
	pattern = re.compile(regex, re.M|re.S|re.DOTALL)
	res = pattern.findall(html)
	data = json.loads(res[0][res[0].find(start):])
	return data

#TODO
def printDictList(pro_city_data, level):
	if(isinstance(pro_city_data,dict)):
		for key in pro_city_data:
			for t in range(1,level): #前闭后开
				print("\t",end=' ')
			print(key,":")
			printDictList(pro_city_data[key],level+1)
	elif(isinstance(pro_city_data,list)):
		for x in pro_city_data:
			printDictList(x,level+1)
	else:
		print(pro_city_data)

#城市属于的省份. 如果不知所属的省份，返回Space
def cityOfProvince(city, china_city):
	#china_city
	citylist = china_city['citylist'] # 列表

	for province in citylist: # 每个province是一个dict
		province_name = province.get('p','TTTT')
		if(city in province_name):
			return province_name
		cityInpro = province.get('c',[]) #是一个list
		for one in cityInpro: #one 是一个dict
			if(city in one.get('n','space')):
				return province_name
			for quxian in one.get('a',[]):
				if(city in quxian.get('s','')):
					return province_name
	return "Space"

def getShopNum(city):
	city = urllib.parse.quote(city.encode('utf-8'))
	url = 'http://www.haidilao.com/service/ajax_store?city='+city 
	html = getHtml(url)
	data = json.loads(html)
	#print(data)
	#print(len(data))
	return len(data)

def staticsShop(china_city, shop_city):
	#statics_data = {} # {"福建":{"num":12,"city":[{"福州":10},{"厦门":10}]}}
	
	#shop_city是一个list
	for city in shop_city:
		cityname = city[:city.find('|')]
		#print('cityname:' + cityname)
		shop_num = getShopNum(cityname)
		#print(shop_num)
		citydict = {}
		citydict[cityname] = shop_num

		province = cityOfProvince(cityname,china_city)

		#print(province)
		if(not statics_data.get(province)):
			statics_data[province] = {}
		if(not statics_data[province].get('num')):
			statics_data[province]['num'] = 0
		statics_data[province]['num'] += shop_num
		if(not statics_data[province].get('city')):
			statics_data[province]['city'] = []
		statics_data[province]['city'].append(citydict)

	result = sorted(statics_data.items(), key = lambda x: x[1].get("num"), reverse=True)

	exportExcel(result)

	with open(data_name,"w+") as f:
		f.write(json.dumps(result,ensure_ascii=False, indent=4, separators=(',', ': ')))

	#print(statics_data)

#利用多线程
def staticsShop_thread(china_city, shop_city):

	threads = []
	#开10线程
	for i in range(1,10):
		t = threading.Thread(target=process_thread, args=("Thread-"+str(i),shop_city,china_city))
		t.start()
		threads.append(t)

	for t in threads:
		t.join()

	result = sorted(statics_data.items(), key = lambda x: x[1].get("num"), reverse=True)
	exportExcel(result)
	with open(data_name,"w+") as f:
		f.write(json.dumps(result,ensure_ascii=False, indent=4, separators=(',', ': ')))

	#print(statics_data)

def old_http_main():
	html = getHtml(china_city_url)
	#print(html)
	#china_city = getCity(html)
	#统一成一个函数
	china_city = getDataFromHtml(html,'pro_city_data.*"国外"}]}' , '{')

	# with open("china_city.txt","w+") as f:
	# 	# ensure 正常显示中文，而不是显示ASCII值。
	# 	f.write(json.dumps(china_city,ensure_ascii=False, indent=4, separators=(',', ': ')))
	
	html = getHtml(haidilao_city_url)
	#shop_city = getHaidilaoCity(html)
	shop_city = getDataFromHtml(html, 'varallCities=.*gx"]', '[')
	#print(shop_city_data)
	# with open("shop_city.txt","w+") as f:
	# 	f.write(json.dumps(shop_city,ensure_ascii=False, indent=4, separators=(',', ': ')))

	#staticsShop(china_city,shop_city)

	staticsShop_thread(china_city, shop_city)

	#sendmail.sendEmail(["1043096262@qq.com"],'海底捞门店','Fighting',['result.xlsx','static_data.txt'])
	sendmail.sendEmail(email_list,'海底捞门店数据','Fighting',[execel_name])
	#email_package.sendmail.sendEmail(email_list,'海底捞门店','Fighting',[execel_name])



# just test
def printPor():
	global china_cities
	#print(china_cities)
	print(china_cities[0])

def new_https_main():

	global china_cities
	with open(china_file, 'r') as f:
		text = f.read()
		china_cities = json.loads(text)

	#printPor()


	taiwan = {
				'name':'台湾',
				'level':1,
				'code':'',
				'child':[
					{
						'name':'市辖区',
						'level':2,
						'code':'',
						'child':None
					}
				]
	}
	hongkong = {
				'name':'香港',
				'level':1,
				'code':'',
				'child':[
					{
						'name':'市辖区',
						'level':2,
						'code':'',
						'child':None
					}
				]

	}
	macao = {
				'name':'澳门',
				'level':1,
				'code':'',
				'child':[
					{
						'name':'市辖区',
						'level':2,
						'code':'',
						'child':None
					}
				]


	}

	china_cities.append(taiwan)
	china_cities.append(hongkong)
	china_cities.append(macao)
	#每个省份开一个线程。
	threads = []
	
	for i in range(len(china_cities)):
		t = threading.Thread(target=process_province_request, args=("Thread-"+str(i) , i))
		t.start()
		threads.append(t)

	for t in threads:
		t.join()


	print(statics_data)
	result = sorted(statics_data.items(), key = lambda x: x[1].get("num"), reverse=True)
	exportExcel(result)
	with open(data_name,"w+") as f:
		f.write(json.dumps(result,ensure_ascii=False, indent=4, separators=(',', ': ')))

	#email_package.sendmail.sendEmail(email_list,'海底捞门店','Fighting',[execel_name])
	sendmail.sendEmail(email_list,'海底捞门店数据','Fighting',[execel_name])



if __name__ == '__main__':
	start = time.time()
	#old_http_main()
	new_https_main()
	end = time.time()
	print(end - start)